"""
export.py
---------
Handles turning a match's data into downloadable files: PDF scorecards
(reportlab), Excel workbooks (openpyxl via pandas), and plain CSV.
Every function returns bytes so Streamlit's st.download_button can
serve them directly without leaving stray temp files, while also
saving a copy under exports/<format>/ for the "Previous Matches" page.
"""

import io
import os
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                 Paragraph, Spacer)
from reportlab.lib.units import cm

import config
import utils


def _innings_dataframes(db, innings_id):
    batting = db.get_batting_card(innings_id)
    bowling = db.get_bowling_card(innings_id)

    bat_rows = []
    for r in batting:
        bat_rows.append({
            "Batter": r["player_name"],
            "Status": r["dismissal_type"] if r["is_out"] else r["status"],
            "Runs": r["runs"], "Balls": r["balls"],
            "4s": r["fours"], "6s": r["sixes"],
            "SR": utils.calc_strike_rate(r["runs"], r["balls"]),
        })
    bowl_rows = []
    for r in bowling:
        bowl_rows.append({
            "Bowler": r["player_name"],
            "Overs": utils.balls_to_overs_str(r["balls"]),
            "Maidens": r["maidens"], "Runs": r["runs_conceded"],
            "Wickets": r["wickets"],
            "Economy": utils.calc_economy(r["runs_conceded"], r["balls"]),
        })
    return pd.DataFrame(bat_rows), pd.DataFrame(bowl_rows)


# --------------------------------------------------------------------------
# CSV EXPORT
# --------------------------------------------------------------------------
def export_csv(db, match_id):
    match = db.get_match(match_id)
    innings_list = db.get_innings_by_match(match_id)

    buffer = io.StringIO()
    for idx, inn in enumerate(innings_list, start=1):
        bat_df, bowl_df = _innings_dataframes(db, inn["innings_id"])
        team_name = db.get_team_name(inn["batting_team_id"])
        buffer.write(f"Innings {idx} - {team_name}\n")
        buffer.write("BATTING\n")
        bat_df.to_csv(buffer, index=False)
        buffer.write("\nBOWLING\n")
        bowl_df.to_csv(buffer, index=False)
        buffer.write("\n\n")

    data = buffer.getvalue().encode("utf-8")
    filename = f"{match['match_name'].replace(' ', '_')}_scorecard.csv"
    path = os.path.join(config.EXPORT_CSV_DIR, filename)
    with open(path, "wb") as f:
        f.write(data)
    return data, filename


# --------------------------------------------------------------------------
# EXCEL EXPORT
# --------------------------------------------------------------------------
def export_excel(db, match_id):
    match = db.get_match(match_id)
    innings_list = db.get_innings_by_match(match_id)

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        for idx, inn in enumerate(innings_list, start=1):
            bat_df, bowl_df = _innings_dataframes(db, inn["innings_id"])
            team_name = db.get_team_name(inn["batting_team_id"])
            sheet_bat = f"Inn{idx}_Batting"[:31]
            sheet_bowl = f"Inn{idx}_Bowling"[:31]
            bat_df.to_excel(writer, sheet_name=sheet_bat, index=False)
            bowl_df.to_excel(writer, sheet_name=sheet_bowl, index=False)

            for sheet_name in (sheet_bat, sheet_bowl):
                ws = writer.sheets[sheet_name]
                header_fill = PatternFill(start_color="1B2030", end_color="1B2030",
                                           fill_type="solid")
                for cell in ws[1]:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
                for col in ws.columns:
                    max_len = max((len(str(c.value)) for c in col if c.value is not None),
                                  default=10)
                    ws.column_dimensions[col[0].column_letter].width = max_len + 4

    data = buffer.getvalue()
    filename = f"{match['match_name'].replace(' ', '_')}_scorecard.xlsx"
    path = os.path.join(config.EXPORT_EXCEL_DIR, filename)
    with open(path, "wb") as f:
        f.write(data)
    return data, filename


# --------------------------------------------------------------------------
# PDF EXPORT
# --------------------------------------------------------------------------
def export_pdf(db, match_id):
    match = db.get_match(match_id)
    innings_list = db.get_innings_by_match(match_id)
    result = db.get_result(match_id)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                             topMargin=1.5 * cm, bottomMargin=1.5 * cm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("TitleStyle", parent=styles["Title"],
                                  textColor=colors.HexColor("#1B2030"))
    heading_style = ParagraphStyle("HeadingStyle", parent=styles["Heading2"],
                                    textColor=colors.HexColor("#00A88F"))

    elements = [
        Paragraph(match["match_name"] or "Cricket Match", title_style),
        Paragraph(f"{match.get('tournament_name') or ''} | {match.get('venue') or ''} | "
                  f"{match.get('match_date') or ''}", styles["Normal"]),
        Spacer(1, 12),
    ]

    for idx, inn in enumerate(innings_list, start=1):
        bat_df, bowl_df = _innings_dataframes(db, inn["innings_id"])
        team_name = db.get_team_name(inn["batting_team_id"])
        elements.append(Paragraph(
            f"Innings {idx}: {team_name} - {inn['total_runs']}/{inn['total_wickets']} "
            f"({utils.balls_to_overs_str(inn['total_balls'])} overs)", heading_style))
        elements.append(Spacer(1, 6))

        if not bat_df.empty:
            bat_table_data = [list(bat_df.columns)] + bat_df.astype(str).values.tolist()
            t = Table(bat_table_data, hAlign="LEFT")
            t.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1B2030")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
            ]))
            elements.append(t)
            elements.append(Spacer(1, 10))

        if not bowl_df.empty:
            bowl_table_data = [list(bowl_df.columns)] + bowl_df.astype(str).values.tolist()
            t2 = Table(bowl_table_data, hAlign="LEFT")
            t2.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1B2030")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
            ]))
            elements.append(t2)
        elements.append(Spacer(1, 16))

    if result:
        elements.append(Paragraph("Result", heading_style))
        elements.append(Paragraph(result["summary"] or "", styles["Normal"]))

    doc.build(elements)
    data = buffer.getvalue()
    filename = f"{match['match_name'].replace(' ', '_')}_scorecard.pdf"
    path = os.path.join(config.EXPORT_PDF_DIR, filename)
    with open(path, "wb") as f:
        f.write(data)
    return data, filename
